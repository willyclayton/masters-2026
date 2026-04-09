#!/usr/bin/env python3
"""
Parse masters_bets.xlsx into data/bets.json.

The xlsx is a manual log; this script converts each row's free-text Selection
string into typed legs (player + market + per-leg American odds) so the Next.js
app can render a Bets Tracking tab without parsing xlsx at runtime.

Usage:
    python3 scripts/parse-bets.py

Re-run any time bets are added/edited in the xlsx.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "masters_bets.xlsx"
PLAYERS = ROOT / "data" / "players-2026.json"
OUT = ROOT / "data" / "bets.json"

# Market phrase → (key, label) -- ordered longest first to avoid prefix collisions
MARKET_PATTERNS = [
    (re.compile(r"top\s*former\s*masters\s*winner", re.I), "topFormer", "Top Former Masters Winner"),
    (re.compile(r"top\s*debutant\s*golfer", re.I),         "topDebutant", "Top Debutant Golfer"),
    (re.compile(r"top\s*asian\s*golfer", re.I),            "topAsian", "Top Asian Golfer"),
    (re.compile(r"top\s*american\s*golfer", re.I),         "topAmerican", "Top American Golfer"),
    (re.compile(r"top\s*european\s*golfer", re.I),         "topEuropean", "Top European Golfer"),
    (re.compile(r"tournament\s*winner", re.I),             "win", "Tournament Winner"),
    (re.compile(r"top\s*5\s*finish", re.I),                "top5", "Top 5 Finish"),
    (re.compile(r"top\s*10\s*finish", re.I),               "top10", "Top 10 Finish"),
    (re.compile(r"top\s*20\s*finish", re.I),               "top20", "Top 20 Finish"),
    (re.compile(r"\bT5\b"),                                 "top5", "Top 5 Finish"),
    (re.compile(r"\bT10\b"),                                "top10", "Top 10 Finish"),
    (re.compile(r"\bT20\b"),                                "top20", "Top 20 Finish"),
    (re.compile(r"make\s*the\s*cut", re.I),                "makeCut", "Make The Cut"),
]

ODDS_RE = re.compile(r"\(([+-]\d+)\)")


def load_player_names():
    return {p["name"] for p in json.loads(PLAYERS.read_text())}


def detect_market(text: str):
    for pat, key, label in MARKET_PATTERNS:
        m = pat.search(text)
        if m:
            return key, label, m
    return None, None, None


def find_player(text: str, players) -> Optional[str]:
    """Return the longest player name that appears as a substring of `text`."""
    matches = [p for p in players if p in text]
    if not matches:
        return None
    return max(matches, key=len)


def parse_leg(text: str, players: set[str], default_market: tuple | None = None):
    """Parse a single leg fragment into {player, market, marketLabel, americanOdds}."""
    text = text.strip()

    odds_match = ODDS_RE.search(text)
    american = odds_match.group(1) if odds_match else None
    text_no_odds = ODDS_RE.sub("", text).strip()

    if default_market is not None:
        market_key, market_label = default_market
        residual = text_no_odds
    else:
        market_key, market_label, m = detect_market(text_no_odds)
        if not market_key:
            raise ValueError(f"Could not detect market in leg: {text!r}")
        residual = (text_no_odds[: m.start()] + " " + text_no_odds[m.end() :]).strip(" -")

    player = find_player(residual, players)
    if not player:
        raise ValueError(f"Could not match a player name in leg: {text!r} (residual={residual!r})")

    return {
        "player": player,
        "market": market_key,
        "marketLabel": market_label,
        "americanOdds": american,
    }


def parse_selection(selection: str, bet_type: str, players: set[str]):
    """Parse the Selection cell into a list of legs."""
    selection = selection.strip()

    # "Make The Cut: A (-400) / B (-800) / ..." style
    mc_prefix = re.match(r"make\s*the\s*cut\s*:\s*", selection, re.I)
    if mc_prefix:
        body = selection[mc_prefix.end():]
        chunks = [c.strip() for c in body.split("/") if c.strip()]
        return [parse_leg(c, players, default_market=("makeCut", "Make The Cut")) for c in chunks]

    if "/" in selection:
        chunks = [c.strip() for c in selection.split("/") if c.strip()]
        return [parse_leg(c, players) for c in chunks]

    # Straight: "Player Name - Market Phrase"
    if " - " in selection:
        player_part, market_part = selection.split(" - ", 1)
        market_key, market_label, _ = detect_market(market_part)
        if not market_key:
            raise ValueError(f"Could not detect market in straight: {selection!r}")
        if player_part.strip() not in players:
            # try substring match as a fallback
            player = find_player(player_part, players)
            if not player:
                raise ValueError(f"Player not found: {player_part!r}")
        else:
            player = player_part.strip()
        return [{
            "player": player,
            "market": market_key,
            "marketLabel": market_label,
            "americanOdds": None,  # straight uses bet-level odds
        }]

    raise ValueError(f"Unrecognized selection format: {selection!r}")


def main() -> int:
    players = load_player_names()
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    bets = []
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["#"]] or not isinstance(row[idx["#"]], int):
            continue  # skip totals row

        bet_id = row[idx["#"]]
        bet_type = row[idx["Type"]]
        selection = row[idx["Selection"]] or ""
        odds = row[idx["Odds"]]
        stake = row[idx["Bet"]]
        payout = row[idx["Potential Payout"]]
        placed = row[idx["Placed"]]

        try:
            legs = parse_selection(selection, bet_type, players)
        except ValueError as e:
            errors.append(f"Bet #{bet_id}: {e}")
            continue

        # For straights with no per-leg odds, copy bet-level odds onto the single leg.
        if bet_type == "Straight" and len(legs) == 1 and legs[0]["americanOdds"] is None:
            legs[0]["americanOdds"] = odds

        bets.append({
            "id": bet_id,
            "type": bet_type,
            "stake": stake,
            "potentialPayout": payout,
            "combinedAmericanOdds": odds,
            "placedAt": placed,
            "legs": legs,
        })

    if errors:
        print("PARSE ERRORS:", file=sys.stderr)
        for e in errors:
            print("  " + e, file=sys.stderr)
        return 1

    out = {
        "lastUpdated": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "source": "masters_bets.xlsx",
        "bets": bets,
    }
    OUT.write_text(json.dumps(out, indent=2) + "\n")

    print(f"Wrote {OUT.relative_to(ROOT)} ({len(bets)} bets)")
    for b in bets:
        legs_desc = ", ".join(f"{l['player']} {l['market']}" for l in b["legs"])
        print(f"  #{b['id']:>2} {b['type']:8} {b['combinedAmericanOdds']:>7}  ${b['stake']}→${b['potentialPayout']}  {legs_desc}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
